import openpyxl
from openpyxl.styles import PatternFill
import tkinter as tk
'''
筛选规则
全部输出
     重复率小于50，        不填充        0
     重复率大于50，连续     单元格标黄     1
                 不连续   单元格标红     2 
'''


def load_dataset(url):
    data_set = []
    wb = openpyxl.load_workbook(url)
    ws = wb.worksheets[0]

    # 数据库在第几列
    B_col = ws['B']
    for each in B_col:
        data_set.append(each.value)
    # 排除了第一行的表头
    return data_set[1:]     # 输出为列表


def load_inputset(url):
    input_set = []
    wb = openpyxl.load_workbook(url)
    ws = wb.worksheets[0]

    # 筛选目标在第几列
    A_col = ws['A']
    for each in A_col:
        input_set.append(each.value)
    # 排除了第一行的表头
    return input_set[1:]    # 输出为列表


def is_continuous(lengh, each_input):
    """
    重复率小于50，        不填充        标记0
    重复率大于50，连续     单元格标黄    标记1
                不连续    单元格标红   标记2
    """
    l1 = []
    aa = sorted(lengh)
    lis = []
    for x in sorted(set(aa)):
        l1.append(x)
        if x + 1 not in aa:
            if len(l1) != 1:
                lis.append(l1)
            l1 = []

    if len(lengh) >= (len(each_input) / 2):     # 重复率大于50
        for each_li in lis:
            if len(each_li) >= (len(each_input) / 2):  # 连续
                return 1
            else:   # 不连续
                return 2
    else:   # 重复率小于50
        return 0


# 输出为[筛选目标, 结果1, 结果2 ....]
def compare_tool(each_input, dataset):
    value = []
    value.append(each_input)
    for each in dataset:
        if each in each_input:
            value.append(each)

    lengh = set()
    for each in dataset:
        bag = 0
        if not each_input.find(each) == -1:     # 判断是否重复
            while not bag == -1:
                for i in range(len(each)):
                    lengh.add(bag + i)          # 记录重复的index
                bag = each_input.find(each, bag+1)  # 全句搜索

    flag = is_continuous(lengh, each_input)     # 按填充规则判断
    value.insert(0, flag)
    return value


def save(datalist, save_url, save_filename, text, v):
    v.set('保存中，请稍后.....')
    fill0 = PatternFill('solid')
    fill1 = PatternFill('solid', fgColor='FFFF00')
    fill2 = PatternFill('solid', fgColor='FF0000')
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('结果', 0)
    # 写入表头
    ws.cell(1, 1, '筛选目标')
    ws.cell(1, 2, '结果')
    # 写入数据
    row = 2
    for i in datalist:
        ws.append(i[1:])

        # 填充
        if i[0] == 0:
            ws.cell(row, 1).fill = fill0
        elif i[0] == 1:
            ws.cell(row, 1).fill = fill1
        elif i[0] == 2:
            ws.cell(row, 1).fill = fill2
        row += 1

    wb.save(f'{save_url}\{save_filename}')
    v.set('保存完毕')
    text.insert(tk.END, f'结果已保存至<{save_filename}>文件')


def tool5_main(dataset_url, input_url, save_url, save_filename, text, v):

    # 输入文件所在位置
    dataset_url = dataset_url.get()
    input_url = input_url.get()
    save_url = save_url.get()
    save_filename = save_filename.get()

    # 读取数据库和筛选目标
    dataset = load_dataset(dataset_url)
    inputset = load_inputset(input_url)

    # 测试数据（无视）
    # inputset = ['ABCDEFG', 'ABCD', 'EDEDPOIU', 'EDEDPOILKJ', 'PRIACKK', 'AFK']
    # dataset = ['AB', 'CD', 'BCD', 'ABD', 'ABCD', 'ABCDE', 'EDE', 'DED', 'PR', 'KK']
    # inputset = ['EEEEQRQ']
    # dataset = ['E', 'EE', 'EEE']


    # 筛选目标
    count = 1
    outputlist = []
    total = len(inputset)
    v.set(f'共有{total}个待处理')
    for each_input in inputset:
        output = compare_tool(each_input, dataset)
        # print(output)
        outputlist.append(output)
        text.insert(tk.END, f'正在处理第{count}/{total}个，{each_input}\n')
        text.see(tk.END)
        count += 1
    # 保存
    save(outputlist, save_url, save_filename, text, v)
    v.set('筛选已完成')


if __name__ == '__main__':
    tool5_main()
