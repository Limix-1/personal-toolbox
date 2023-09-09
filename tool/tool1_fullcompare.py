import openpyxl


# 筛选
# 长度大于min，小于max的连续字符组合
def tool(s, min, max):
    res = []
    for i in range(min, len(s)+1):
        for j in range(len(s)-i+1):
            if len(s[j:j+i]) <= max:
                if s[j:j+i] not in res:
                    res.append(s[j:j+i])
    res = list(sorted(res, key=lambda x: (len(x), x)))
    return res


# 保存到文件
def save(all, all_re, save_directory, var, filename):
    var.set('正在保存中')
    column = 1
    row = 1
    # 把结果写入新的表格
    wb2 = openpyxl.Workbook()
    ws2_1 = wb2.create_sheet('去重后', 0)
    ws2_2 = wb2.create_sheet('去重前', 1)


    # 保存去重后
    for val in all_re:
        ws2_1.cell(row, column, val)
        row += 1

    row = 1
    # 保存去重前
    for val in all:
        ws2_2.cell(row, column, val)
        row += 1

    wb2.save(f'{save_directory}\{filename}')
    var.set(f'去重前有{len(all)}个元素\n去重后有{len(all_re)}个元素\n结果已保存至{save_directory}\{filename}')


def full_com(obj_file_name, save_directory, var, mi, ma, filename):
    obj_file_name = obj_file_name.get()
    save_directory = save_directory.get()
    mi = int(mi.get())
    ma = int(ma.get())
    save_filename = filename.get()
    var.set('工具正在运行中')

    # 打开xlsx文件
    wb1 = openpyxl.load_workbook(obj_file_name)
    # 打开第1个工作簿
    ws1 = wb1.worksheets[0]
    # 打开工作簿的第A列 赋值为a
    a = ws1['A']
    print('一共有{}个元素需要筛选'.format(len(a)))
    all = []
    for i in a:
        # 把所有 值 添加到x
        x = i.value
        # 把返回的过滤后的列表添加到all
        all.extend(tool(x, mi, ma))

    # 排序
    all = list(sorted(all, key=lambda x: (len(x), x)))

    # 去重排序  # 用集合去重
    all_re = list(sorted(list(set(all)), key=lambda x: (len(x), x)))

    # 保存
    save(all, all_re, save_directory, var, save_filename)


if __name__ == '__main__':
    obj_file_name = "/tool_1/新建 XLSX 工作表.xlsx"
    save_directory = "D:\Desktop\脚本文件\\tool_1"
    mi = 3
    ma = 3
    var = '1'
    filename = "结果.xlsx"
    # full_com(obj_file_name, save_directory, var, mi, ma, filename)
