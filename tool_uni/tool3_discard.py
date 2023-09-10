import openpyxl


def load_dataset(url):
    data_set = []
    wb = openpyxl.load_workbook(url)
    ws = wb.worksheets[0]

    # 数据库在第几列
    B_col = ws['B']
    for each in B_col:
        data_set.append(each.value)
    # 排除了第一行的表头
    return data_set[1:]     # 输出为列表


def load_inputset(url):
    input_set = []
    wb = openpyxl.load_workbook(url)
    ws = wb.worksheets[0]

    # 筛选目标在第几列
    A_col = ws['A']
    for each in A_col:
        input_set.append(each.value)
    # 排除了第一行的表头
    return input_set[1:]    # 输出为列表


def compare_tool(each_input, dataset):
    value = []
    value.append(each_input)
    for each in dataset:
        if len(each) >= (len(each_input)/2):
            if each in each_input:
                value.append(each)
    # 输出为[筛选目标, 结果1, 结果2 ....]
    return value


def save(datalist):
    print('保存中，请稍后.....')
    row = 1
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('结果', 0)
    # 写入表头
    ws.cell(1, 1, '筛选目标')
    ws.cell(1, 2, '结果')
    # 写入数据
    for i in datalist:
        ws.append(i)
    wb.save('结果1.xlsx')
    print('结果已保存至<结果1.xlsx>文件')


if __name__ == '__main__':

    # 输入文件所在位置
    dataset_url = r'D:\Desktop\新建文件夹\导入文件-鲜味数据库.xlsx'
    input_url = r'D:\Desktop\新建文件夹\检索文件.xlsx'

    # 读取数据库和筛选目标
    dataset = load_dataset(dataset_url)
    inputset = load_inputset(input_url)

    # inputset = ['ABCDEFG', 'ABCD']
    # dataset = ['AB', 'CD', 'BCD', 'ABD', 'ABCD', 'ABCDE']

    # 筛选目标
    count = 1
    outputlist = []
    total = len(inputset)
    for each_input in inputset:
        output = compare_tool(each_input, dataset)
        outputlist.append(output)
        print(f'正在处理第{count}/{total}个，{each_input}')
        count += 1

    # 保存
    save(outputlist)
    print('筛选已完成')
