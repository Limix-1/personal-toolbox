import openpyxl
import tkinter as tk
'''
输出规则
# 重复小于50       不输出
# 重复大于50       输出
'''



def load_dataset(url):
    data_set = []
    wb = openpyxl.load_workbook(url)
    ws = wb.worksheets[0]

    # 数据库在第几列
    B_col = ws['B']
    for each in B_col:
        data_set.append(each.value)
    # 排除了第一行的表头
    return data_set[1:]     # 输出为列表


def load_inputset(url):
    input_set = []
    wb = openpyxl.load_workbook(url)
    ws = wb.worksheets[0]

    # 筛选目标在第几列
    A_col = ws['A']
    for each in A_col:
        input_set.append(each.value)
    # 排除了第一行的表头
    return input_set[1:]    # 输出为列表


    # 输出为[筛选目标, 结果1, 结果2 ....]
def compare_tool(each_input, dataset):
    value = []
    value.append(each_input)
    lengh = set()
    for each in dataset:
        bag = 0
        if not each_input.find(each) == -1:
            value.append(each)
            while not bag == -1:
                for i in range(len(each)):
                    lengh.add(bag + i)
                bag = each_input.find(each, bag+1)
    if len(lengh) >= (len(each_input) / 2):
        return value
    else:
        return [each_input]


def save(datalist, save_url, save_filename, text, v):
    v.set('保存中，请稍后.....')
    row = 1
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('结果', 0)
    # 写入表头
    ws.cell(1, 1, '筛选目标')
    ws.cell(1, 2, '结果')
    # 写入数据
    for i in datalist:
        ws.append(i)
    wb.save(f'{save_url}\{save_filename}')
    v.set('保存完毕')
    text.insert(tk.END, f'结果已保存至<{save_filename}>文件')


def too4_main(dataset_url, input_url, save_url, save_filename, text, v):
    # 输入文件所在位置
    dataset_url = dataset_url.get()
    input_url = input_url.get()
    save_url = save_url.get()
    save_filename = save_filename.get()

    # 读取数据库和筛选目标
    dataset = load_dataset(dataset_url)
    inputset = load_inputset(input_url)

    # 筛选目标
    count = 1
    outputlist = []
    total = len(inputset)
    for each_input in inputset:
        output = compare_tool(each_input, dataset)
        # print(output)
        outputlist.append(output)
        text.insert(tk.END, f'正在处理第{count}/{total}个，{each_input}\n')
        count += 1
    # 保存
    save(outputlist, save_url, save_filename, text, v)
    v.set('筛选已完成')


if __name__ == '__main__':
    too4_main()

